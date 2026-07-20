"""Exporta la lista de actividades diarias a Excel (.xlsx).

Pensado para auditoria SCJ: el casino necesita entregar un Excel
con las mismas 11 columnas del archivo que antes se llenaba a mano.

Columnas (orden del Excel original):
  1. Fecha         (dd/mm/aaaa)
  2. Produccion / Tarea
  3. Area
  4. Maquina
  5. Detalle
  6. Isla
  7. Ticket Jira (si/no)
  8. Numero de Ticket de Jira
  9. Pendiente (si/no)
  10. Tecnico
  11. Turno
"""

from __future__ import annotations

from datetime import date
from pathlib import Path


# Mapeo de bool -> "si"/"no" para que el Excel abra igual al original.
def _sn(b: bool) -> str:
    return "si" if b else "no"


def _fmt_fecha(s: str | None) -> str:
    """Acepta ISO (yyyy-mm-dd) o date, devuelve dd/mm/aaaa."""
    if not s:
        return ""
    if isinstance(s, date):
        return s.strftime("%d/%m/%Y")
    # Viene como ISO string
    try:
        d = date.fromisoformat(s[:10])
        return d.strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return str(s)


# Encabezados en espanol tal como los espera la SCJ.
HEADERS: list[str] = [
    "Fecha",
    "Produccion",  # columna "Tarea" del Excel original
    "Area",
    "Maquina",
    "Detalle",
    "Isla",
    "Ticket Jira",
    "Numero de Ticket de Jira",
    "Pendiente",
    "Tecnico",
    "Turno",
]


def exportar_excel(
    registros: list[dict],
    destino: Path,
    *,
    titulo: str = "Registro de Actividades Diarias",
) -> int:
    """Escribe los registros al .xlsx en ``destino``.

    Retorna la cantidad de filas exportadas (sin contar el header).
    Lanza ``OSError`` si no se puede escribir.

    El archivo se genera con:
      - Header en negrita + fondo gris
      - Ancho de columna ajustado al contenido
      - Una sola hoja con el titulo como nombre
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as ex:
        raise RuntimeError(
            "openpyxl no esta instalado. Agregalo a requirements.txt "
            "y reinstala la app."
        ) from ex

    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Excel limita el nombre de la hoja a 31 chars

    # Estilos del header
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill(start_color="1E5AA8", end_color="1E5AA8",
                            fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Header
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = wrap

    # Filas
    for row_idx, r in enumerate(registros, start=2):
        ws.cell(row=row_idx, column=1, value=_fmt_fecha(r.get("fecha")))
        ws.cell(row=row_idx, column=2, value=str(r.get("tarea") or ""))
        ws.cell(row=row_idx, column=3, value=str(r.get("area") or ""))
        ws.cell(row=row_idx, column=4, value=str(r.get("numero_maquina") or ""))
        ws.cell(row=row_idx, column=5, value=str(r.get("detalle") or ""))
        ws.cell(row=row_idx, column=6, value=str(r.get("isla") or ""))
        ws.cell(row=row_idx, column=7, value=_sn(bool(r.get("ticket_jira_sn"))))
        ws.cell(row=row_idx, column=8, value=str(r.get("numero_ticket_jira") or ""))
        ws.cell(row=row_idx, column=9, value=_sn(bool(r.get("pendiente_sn"))))
        ws.cell(row=row_idx, column=10, value=str(r.get("tecnico") or ""))
        ws.cell(row=row_idx, column=11, value=str(r.get("turno") or ""))
        for col_idx in range(1, 12):
            ws.cell(row=row_idx, column=col_idx).alignment = wrap

    # Ancho de columnas (aprox, basado en los headers)
    anchos = [12, 28, 14, 10, 60, 10, 12, 22, 12, 22, 14]
    for col_idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = ancho

    # Fila 1 un poco mas alta para que respire el header en negrita
    ws.row_dimensions[1].height = 22

    wb.save(destino)
    return len(registros)


__all__ = ("exportar_excel", "HEADERS")
