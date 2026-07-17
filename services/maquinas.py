"""Servicios de maquinas: busqueda, CRUD, importacion desde Excel.

Reglas de negocio:

* ``numero_maquina`` es unico en el catalogo (constraint logico).
* Si se intenta crear una maquina con un numero existente y la
  maquina esta activa, se actualiza el resto de campos (upsert).
* La busqueda es por prefijo sobre ``numero_maquina`` y LIKE sobre
  marca/modelo para los resultados del autocomplete.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from sqlalchemy import or_, select
from sqlalchemy.exc import IntegrityError

from config import (
    ESTADOS_MAQUINA,
    LOGS_DIR,
)
from database.db import get_session
from database.models import Maquina


# ---------------------------------------------------------------------------
# Busqueda
# ---------------------------------------------------------------------------


def buscar_maquinas(query: str, limit: int = 25) -> list[dict]:
    """Busca maquinas por numero (prefijo), marca o modelo.

    Parametros
    ----------
    query:
        Texto libre. Si esta vacio devuelve las primeras ``limit``
        maquinas activas.
    limit:
        Cantidad maxima de resultados.

    Retorna
    -------
    list[dict]
        Cada elemento es ``Maquina.to_dict()``.
    """
    q = (query or "").strip()

    with get_session() as s:
        stmt = select(Maquina).where(Maquina.activo.is_(True))

        if q:
            like = f"%{q}%"
            stmt = stmt.where(
                or_(
                    Maquina.numero_maquina.ilike(f"{q}%"),
                    Maquina.numero_maquina.ilike(like),
                    Maquina.marca.ilike(like),
                    Maquina.modelo.ilike(like),
                )
            )

        stmt = stmt.order_by(Maquina.numero_maquina.asc()).limit(limit)
        return [m.to_dict() for m in s.scalars(stmt)]


def obtener_por_numero(numero_maquina: str, incluir_inactivos: bool = False) -> dict | None:
    """Devuelve la maquina con ese numero, o ``None``.

    Por defecto solo busca maquinas activas (las visibles en el
    buscador principal). Con ``incluir_inactivos=True`` tambien trae
    las de la papelera (util para el editor de Settings).
    """
    if not numero_maquina:
        return None
    with get_session() as s:
        stmt = select(Maquina).where(Maquina.numero_maquina == str(numero_maquina).strip())
        if not incluir_inactivos:
            stmt = stmt.where(Maquina.activo.is_(True))
        m = s.scalars(stmt).first()
        return m.to_dict() if m else None


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------


def crear_maquina(datos: dict) -> dict:
    """Crea una maquina. Si ya existe el numero, actualiza en su lugar.

    Retorna ``Maquina.to_dict()`` con el id asignado.
    """
    payload = _normalizar(datos)
    with get_session() as s:
        existente = s.scalars(
            select(Maquina).where(Maquina.numero_maquina == payload["numero_maquina"])
        ).first()
        if existente is not None:
            return actualizar_maquina(existente.id, datos)

        m = Maquina(**payload)
        s.add(m)
        try:
            s.flush()
        except IntegrityError as e:  # pragma: no cover - carrera rara
            raise ValueError(f"No se pudo crear la maquina: {e}") from e
        return m.to_dict()


def actualizar_maquina(maquina_id: int, datos: dict) -> dict:
    """Actualiza campos editables de una maquina existente."""
    payload = _normalizar(datos, partial=True)
    with get_session() as s:
        m = s.get(Maquina, maquina_id)
        if m is None:
            raise ValueError(f"Maquina id={maquina_id} no existe")
        for k, v in payload.items():
            setattr(m, k, v)
        s.flush()
        return m.to_dict()


def eliminar_maquina(maquina_id: int, soft: bool = True) -> None:
    """Elimina (soft por defecto: marca ``activo=False``)."""
    with get_session() as s:
        m = s.get(Maquina, maquina_id)
        if m is None:
            return
        if soft:
            m.activo = False
        else:
            s.delete(m)


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Importacion desde Excel
# ---------------------------------------------------------------------------

# Columnas del Excel real (Cierre_Turno_V2.xlsm, hoja "Master").
# Los nombres del Excel vienen con saltos de linea y variaciones, asi que
# el matching se hace por normalizacion (lower, sin acentos, sin espacios).
COLUMNAS_EXCEL = {
    # clave normalizada -> clave que el sistema entiende
    "numero": "numero_maquina",
    "codigo casino maquina": "numero_maquina",
    "codigo casino": "numero_maquina",
    "maquina": "numero_maquina",
    "codigo scj": "codigo_scj",
    "ubicacion maquina de azar sector": "sector",
    "sector": "sector",
    "ubicacion maquina de azar isla": "isla",
    "isla": "isla",
    "no de serie maquina de azar gabinete": "serie",
    "serie": "serie",
    "fabricante gabinete": "marca",
    "fabricante": "marca",
    "modelo gabinete": "modelo",
    "modelo": "modelo",
    "nombre de modelo del programa de juego": "denominacion",
    "denominacion": "denominacion",
    # Columnas de estado del Excel (no del catalogo maestro, pero utiles):
    "estado": "_estado_excel",
    "estado diario": "_estado_excel",
    "reparable si/no": "_reparable",
    "reparable si/no": "_reparable",
    "problema": "_problema_historico",
    "repuesto": "_repuesto_historico",
    "accion": "_accion_historica",
}


# Estados validos que pueden venir en el Excel.
ESTADOS_VALIDOS_EXCEL = (
    "Operativa",
    "Fuera de Servicio",
    "Pendiente Repuesto",
    "Espera Servicio Tecnico",
    "En Observacion",
)


def _normalizar_estado(valor: str | None) -> str:
    """Mapea el estado del Excel a uno de los ESTADOS_MAQUINA.

    Si viene vacio, '-', o algo desconocido, devuelve 'Operativa' como
    fallback (es lo mas seguro para una maquina del catalogo maestro).
    """
    if not valor:
        return "Operativa"
    v = str(valor).strip()
    if not v or v == "-":
        return "Operativa"
    # Match exacto (case-insensitive)
    for est in ESTADOS_VALIDOS_EXCEL:
        if v.lower() == est.lower():
            return est
    # Match parcial: si contiene "fuera" -> FDS, "pendiente" -> Pend Rep, etc
    vl = v.lower()
    if "fuera" in vl:
        return "Fuera de Servicio"
    if "pendiente" in vl:
        return "Pendiente Repuesto"
    if "espera" in vl:
        return "Espera Servicio Tecnico"
    if "observacion" in vl:
        return "En Observacion"
    return "Operativa"


def importar_desde_excel(
    ruta_archivo: str,
    *,
    hoja: str | None = None,
    incluir_inactivos: bool = True,
) -> dict:
    """Importa maquinas desde un Excel.

    Si el archivo tiene la hoja "Master" con el formato de Cierre_Turno_V2.xlsm,
    se carga desde ahi. Si no, se carga la primera hoja con las columnas
    reconocibles (Numero/Sector/Isla/Marca/Modelo/Serie/Denominacion).

    Comportamiento:

    * Si la maquina existe (por ``numero_maquina``), actualiza campos.
    * Si no existe, la inserta con ``activo=True`` y el estado del Excel.

    Retorna
    -------
    dict con conteo ``insertadas``, ``actualizadas``, ``errores`` y
    ``hoja_usada``.
    """
    from openpyxl import load_workbook

    # Abrimos con data_only=True para tomar valores calculados por formulas
    wb = load_workbook(ruta_archivo, data_only=True, read_only=True)

    # Si no nos dicen hoja, priorizamos "Master", despues "Datos", sino la primera
    hoja_elegida: str | None = hoja
    if hoja_elegida is None:
        for candidata in ("Master", "Datos"):
            if candidata in wb.sheetnames:
                hoja_elegida = candidata
                break
    if hoja_elegida is None:
        # Tomar la primera hoja que tenga las columnas reconocibles
        for sn in wb.sheetnames:
            ws_temp = wb[sn]
            primera_fila = next(ws_temp.iter_rows(min_row=1, max_row=1, values_only=True), ())
            if primera_fila and any(
                _norm_col(str(c or "")) in COLUMNAS_EXCEL
                for c in primera_fila
            ):
                hoja_elegida = sn
                break
    if hoja_elegida is None:
        raise ValueError(
            f"No se encontro una hoja con columnas reconocibles en {ruta_archivo}. "
            f"Hojas disponibles: {wb.sheetnames}"
        )

    ws = wb[hoja_elegida]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()

    if not filas:
        return {"insertadas": 0, "actualizadas": 0, "errores": ["archivo vacio"], "hoja_usada": hoja_elegida}

    # La primera fila suele tener headers multi-linea en este Excel
    headers_raw = filas[0]
    headers_norm = [_norm_col(str(c or "")) for c in headers_raw]
    # En "Master" la fila 1 son los headers; si estan vacios, buscar en fila 2 o 3
    if not any(headers_norm):
        for offset in (1, 2):
            if len(filas) > offset:
                cand = [_norm_col(str(c or "")) for c in filas[offset]]
                if any(cand):
                    headers_raw = filas[offset]
                    headers_norm = cand
                    data_start = offset + 1
                    break
        else:
            return {"insertadas": 0, "actualizadas": 0, "errores": ["no se encontraron headers"], "hoja_usada": hoja_elegida}
    else:
        data_start = 1

    # Mapear columna indice -> campo del sistema
    col_map: dict[int, str] = {}
    for idx, header in enumerate(headers_norm):
        if header in COLUMNAS_EXCEL:
            col_map[idx] = COLUMNAS_EXCEL[header]

    if "numero_maquina" not in col_map.values():
        raise ValueError(
            f"No se encontro la columna de numero de maquina en la hoja "
            f"'{hoja_elegida}'. Headers encontrados: {headers_norm}"
        )

    insertadas = 0
    actualizadas = 0
    errores: list[str] = []

    with get_session() as s:
        for idx, row in enumerate(filas[data_start:], start=data_start + 1):
            try:
                # Construir payload solo con campos que existen en este Excel
                payload: dict = {}
                estado_excel = None
                for col_idx, campo in col_map.items():
                    if col_idx >= len(row):
                        continue
                    valor = row[col_idx]
                    valor_str = _safe_str(valor)
                    if campo.startswith("_"):
                        # Campos meta del Excel (estado, problema historico, etc)
                        if campo == "_estado_excel":
                            estado_excel = _normalizar_estado(valor_str)
                        # los demas (_problema_historico, etc) se descartan
                        continue
                    payload[campo] = valor_str
                if not payload.get("numero_maquina"):
                    errores.append(f"Fila {idx + 1}: numero vacio")
                    continue

                # Si el Excel trae estado, lo usamos. Si no, default Operativa.
                estado_final = estado_excel if estado_excel else "Operativa"
                payload["estado"] = estado_final
                payload["activo"] = True

                # Upsert
                existente = s.scalars(
                    select(Maquina).where(
                        Maquina.numero_maquina == payload["numero_maquina"]
                    )
                ).first()
                if existente is None:
                    s.add(Maquina(**payload))
                    insertadas += 1
                else:
                    for k, v in payload.items():
                        if k == "numero_maquina":
                            continue
                        setattr(existente, k, v)
                    actualizadas += 1
            except Exception as e:
                errores.append(f"Fila {idx + 1}: {e}")
        s.flush()

    # Log simple del resultado
    try:
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        log_path = LOGS_DIR / "importaciones.log"
        with log_path.open("a", encoding="utf-8") as f:
            f.write(
                f"{datetime.now().isoformat()} archivo={ruta_archivo} "
                f"hoja={hoja_elegida} insertadas={insertadas} "
                f"actualizadas={actualizadas} errores={len(errores)}\n"
            )
            for e in errores[:20]:  # solo primeros 20 errores en log
                f.write(f"  - {e}\n")
    except OSError:
        pass

    return {
        "insertadas": insertadas,
        "actualizadas": actualizadas,
        "errores": errores,
        "hoja_usada": hoja_elegida,
    }


# ---------------------------------------------------------------------------
# Plantilla Excel para descarga
# ---------------------------------------------------------------------------

HEADER_MASTER: tuple[str, ...] = (
    "Codigo Casino Maquina",
    "Ubicacion\nMaquina de Azar\nSector",
    "Ubicacion\nMaquina de Azar\nIsla",
    "No de Serie\nMaquina de Azar\nGabinete",
    "Fabricante\nGabinete",
    "Modelo\nGabinete",
    "Nombre de Modelo del Programa de Juego",
    "Estado",
    "Reparable  Si/NO",
    "Problema",
    "Repuesto",
    "Accion",
    "Estado Diario",
)


def generar_plantilla(ruta_destino: str, *, hoja: str = "Master") -> str:
    """Genera un Excel plantilla con los headers de la hoja Master.

    Sirve para que el operador baje un archivo vacio con el formato
    esperado, lo complete a mano o lo use como base para su catalogo.

    Retorna la ruta absoluta del archivo generado.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    # Encabezados en la primera fila
    for col_idx, header in enumerate(HEADER_MASTER, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    # Anchos sugeridos para que se vea bien al abrirlo
    anchos = {
        1: 22,  # Codigo Casino
        2: 22,  # Sector
        3: 12,  # Isla
        4: 26,  # Serie
        5: 36,  # Fabricante
        6: 26,  # Modelo
        7: 36,  # Denominacion / Juego
        8: 22,  # Estado
        9: 18,  # Reparable
        10: 38, # Problema
        11: 22, # Repuesto
        12: 38, # Accion
        13: 22, # Estado Diario
    }
    for col_idx, ancho in anchos.items():
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = ancho
    # Fila 2 vacia con ejemplo comentado para que el operador sepa el formato
    ejemplo = (
        5001, "TERRAZA", 213, "HXU8023421",
        "ARISTOCRAT Technologies, Inc", "HELIX UPRIGHT",
        "GOLDEN AMULET", "Operativa", "-", "-", "-", "-", "Operativa",
    )
    for col_idx, valor in enumerate(ejemplo, start=1):
        ws.cell(row=2, column=col_idx, value=valor)
    # Comentario en la celda de Codigo
    from openpyxl.comments import Comment
    ws.cell(row=1, column=1).comment = Comment(
        "Identificador principal del casino. Unico por maquina.\n"
        "Ejemplos: 5001, 5002, 5238.",
        "Sistema FDS",
    )
    ws.cell(row=1, column=8).comment = Comment(
        "Estados validos: Operativa, Fuera de Servicio, Pendiente Repuesto, "
        "Espera Servicio Tecnico, En Observacion.",
        "Sistema FDS",
    )

    # Asegurar que la carpeta destino existe
    dest = Path(ruta_destino).resolve()
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest))
    return str(dest)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _normalizar(datos: dict, partial: bool = False) -> dict:
    """Limpia y valida un payload de maquina."""
    if not partial:
        if "numero_maquina" not in datos or not str(datos["numero_maquina"]).strip():
            raise ValueError("numero_maquina es obligatorio")
    out: dict = {
        "sector": (datos.get("sector") or "").strip() or None,
        "isla": (datos.get("isla") or "").strip() or None,
        "codigo_scj": (datos.get("codigo_scj") or "").strip() or None,
        "marca": (datos.get("marca") or "").strip() or None,
        "modelo": (datos.get("modelo") or "").strip() or None,
        "serie": (datos.get("serie") or "").strip() or None,
        "denominacion": (datos.get("denominacion") or "").strip() or None,
    }
    if not partial:
        out["numero_maquina"] = str(datos["numero_maquina"]).strip()
    estado = (datos.get("estado") or "Operativa").strip()
    if estado not in ESTADOS_MAQUINA:
        estado = "Operativa"
    out["estado"] = estado
    if "activo" in datos:
        out["activo"] = bool(datos["activo"])
    elif not partial:
        out["activo"] = True
    return out


def _safe_str(v) -> str:  # type: ignore[no-untyped-def]
    """Convierte cualquier valor de openpyxl a string limpio.

    None, NaN y floats NaN se vuelven "".
    """
    if v is None:
        return ""
    # Numeros: int/float normales -> str(int) si es entero
    if isinstance(v, (int,)):
        return str(v)
    if isinstance(v, float):
        if v != v:  # NaN
            return ""
        if v.is_integer():
            return str(int(v))
        return str(v).strip()
    # Cualquier otro caso (string, datetime, etc)
    s = str(v).strip()
    if s.lower() in ("nan", "none", "null"):
        return ""
    return s


def _norm_col(name: str) -> str:
    """Normaliza nombre de columna: lower, sin tildes, sin espacios."""
    s = str(name).strip().lower()
    repl = str.maketrans("áéíóúüñ", "aeiouun")
    s = s.translate(repl)
    return " ".join(s.split())


def listar_activas(limit: int = 500) -> list[dict]:
    """Atajo para listar todas las maquinas activas."""
    return buscar_maquinas("", limit=limit)


__all__: Iterable[str] = (
    "buscar_maquinas",
    "obtener_por_numero",
    "crear_maquina",
    "actualizar_maquina",
    "eliminar_maquina",
    "importar_desde_excel",
    "listar_activas",
    "COLUMNAS_EXCEL",
    "ESTADOS_VALIDOS_EXCEL",
    "_normalizar_estado",
    "generar_plantilla",
    "HEADER_MASTER",
)