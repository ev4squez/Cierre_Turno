"""Dialogo del Asistente de Importacion de Maquinas desde Excel.

Flujo:
  1. Boton 'Descargar plantilla' -> genera un .xlsx con los headers
     correctos en la ruta que elija el operador.
  2. Boton 'Seleccionar archivo' -> file picker.
  3. Boton 'Vista previa' -> lee las primeras 20 filas del Excel y las
     muestra en una tabla dentro del dialogo.
  4. Boton 'Importar' -> ejecuta el importador real (services.maquinas).
     Muestra un resumen al final.

El dialogo emite ``finished`` con un dict de resultado al cerrarse,
incluyendo ``insertadas`` y ``actualizadas`` para que el controller
refresque la UI principal.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from services import maquinas as svc_maq


PREVIEW_ROWS = 20


class ImportDialog(QDialog):
    """Asistente visual para importar maquinas desde Excel."""

    finished_with_result = Signal(dict)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Asistente de Importacion - Maquinas")
        self.setMinimumSize(820, 580)
        self.setModal(True)

        self._ruta_archivo: str | None = None
        self._preview_headers: list[str] = []
        self._preview_data: list[tuple] = []

        self._build_ui()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 20)
        root.setSpacing(14)

        # --- Encabezado ---
        head = QFrame()
        hl = QHBoxLayout(head)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.setSpacing(12)
        icon = QLabel()
        icon.setText("EXCEL")
        icon.setStyleSheet(
            "background:#1E5AA8; color:white; font-weight:700; "
            "padding:8px 12px; border-radius:8px;"
        )
        title = QLabel("Asistente de importacion de maquinas")
        title.setStyleSheet("font-size:16px; font-weight:700; color:#1B2430;")
        sub = QLabel(
            "Carga el catalogo desde un Excel (.xlsx o .xlsm). "
            "El archivo debe tener una hoja 'Master' o equivalente con las "
            "columnas reconocibles (Codigo Casino, Sector, Isla, Fabricante, "
            "Modelo, etc)."
        )
        sub.setWordWrap(True)
        sub.setStyleSheet("color:#64748B; font-size:12px;")
        text_col = QVBoxLayout()
        text_col.setContentsMargins(0, 0, 0, 0)
        text_col.setSpacing(2)
        text_col.addWidget(title)
        text_col.addWidget(sub)
        hl.addWidget(icon)
        hl.addLayout(text_col, 1)
        root.addWidget(head)

        # --- Selector de archivo ---
        file_box = QFrame()
        file_box.setObjectName("panelForm")
        file_box.setStyleSheet(
            "QFrame{background:#FFFFFF; border:1px solid #E1E6ED; border-radius:8px;}"
        )
        fbl = QHBoxLayout(file_box)
        fbl.setContentsMargins(12, 10, 12, 10)
        fbl.setSpacing(8)
        lbl = QLabel("Archivo:")
        lbl.setStyleSheet("font-weight:700; color:#64748B;")
        self._lbl_ruta = QLabel("(ninguno seleccionado)")
        self._lbl_ruta.setStyleSheet(
            "color:#94A3B8; font-style:italic;"
        )
        self._lbl_ruta.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        btn_sel = QPushButton("Seleccionar archivo...")
        btn_sel.setObjectName("btnSecondary")
        btn_sel.setCursor(Qt.PointingHandCursor)
        btn_sel.clicked.connect(self._on_seleccionar)

        btn_plantilla = QPushButton("Descargar plantilla")
        btn_plantilla.setObjectName("btnGhost")
        btn_plantilla.setCursor(Qt.PointingHandCursor)
        btn_plantilla.clicked.connect(self._on_descargar_plantilla)

        fbl.addWidget(lbl)
        fbl.addWidget(self._lbl_ruta, 1)
        fbl.addWidget(btn_sel)
        fbl.addWidget(btn_plantilla)
        root.addWidget(file_box)

        # --- Boton preview + resultado ---
        actions = QFrame()
        al = QHBoxLayout(actions)
        al.setContentsMargins(0, 0, 0, 0)
        al.setSpacing(10)
        self._btn_preview = QPushButton("Ver vista previa (primeras 20 filas)")
        self._btn_preview.setObjectName("btnSecondary")
        self._btn_preview.setCursor(Qt.PointingHandCursor)
        self._btn_preview.clicked.connect(self._on_preview)
        self._btn_preview.setEnabled(False)
        al.addWidget(self._btn_preview)
        al.addStretch(1)
        self._btn_importar = QPushButton("Importar a la base de datos")
        self._btn_importar.setObjectName("btnPrimary")
        self._btn_importar.setCursor(Qt.PointingHandCursor)
        self._btn_importar.clicked.connect(self._on_importar)
        self._btn_importar.setEnabled(False)
        al.addWidget(self._btn_importar)
        root.addWidget(actions)

        # --- Tabla de preview ---
        preview_label = QLabel("Vista previa")
        preview_label.setStyleSheet(
            "color:#64748B; font-weight:700; font-size:11px; "
            "text-transform:uppercase; letter-spacing:1px;"
        )
        root.addWidget(preview_label)

        self._tabla = QTableWidget(0, 0)
        self._tabla.setObjectName("tablaIncidencias")
        self._tabla.setAlternatingRowColors(True)
        self._tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        self._tabla.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._tabla.horizontalHeader().setStretchLastSection(True)
        self._tabla.verticalHeader().setVisible(False)
        self._tabla.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        root.addWidget(self._tabla, 1)

        # --- Barra de progreso + estado ---
        status_row = QFrame()
        sl = QHBoxLayout(status_row)
        sl.setContentsMargins(0, 0, 0, 0)
        sl.setSpacing(10)
        self._progress = QProgressBar()
        self._progress.setRange(0, 0)  # Indeterminado
        self._progress.setVisible(False)
        self._progress.setMaximumHeight(6)
        self._progress.setStyleSheet(
            "QProgressBar{border:none; background:#EEF1F6; border-radius:3px;}"
            "QProgressBar::chunk{background:#1E5AA8; border-radius:3px;}"
        )
        self._lbl_status = QLabel("")
        self._lbl_status.setStyleSheet("color:#64748B; font-size:12px;")
        sl.addWidget(self._progress, 1)
        sl.addWidget(self._lbl_status, 1)
        root.addWidget(status_row)

        # --- Botones cerrar ---
        bottom = QFrame()
        bl = QHBoxLayout(bottom)
        bl.setContentsMargins(0, 0, 0, 0)
        bl.setSpacing(10)
        bl.addStretch(1)
        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.setObjectName("btnSecondary")
        btn_cerrar.setCursor(Qt.PointingHandCursor)
        btn_cerrar.clicked.connect(self.accept)
        bl.addWidget(btn_cerrar)
        root.addWidget(bottom)

    # ------------------------------------------------------------------
    # Handlers
    # ------------------------------------------------------------------

    def _on_seleccionar(self) -> None:
        ruta, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar archivo Excel",
            str(Path.home()),
            "Archivos Excel (*.xlsx *.xlsm);;Todos los archivos (*)",
        )
        if ruta:
            self._ruta_archivo = ruta
            self._lbl_ruta.setText(ruta)
            self._lbl_ruta.setStyleSheet("color:#1B2430; font-weight:600;")
            self._btn_preview.setEnabled(True)
            self._btn_importar.setEnabled(False)  # hasta que vea la preview
            self._lbl_status.setText("Archivo seleccionado. Hacé 'Vista previa' para revisar antes de importar.")

    def _on_descargar_plantilla(self) -> None:
        nombre_defecto = f"plantilla_maquinas_{Path.home().stem}.xlsx"
        ruta, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar plantilla",
            str(Path.home() / "Desktop" / nombre_defecto),
            "Excel (*.xlsx)",
        )
        if not ruta:
            return
        try:
            generada = svc_maq.generar_plantilla(ruta)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo generar la plantilla:\n{e}")
            return
        QMessageBox.information(
            self,
            "Plantilla guardada",
            f"Plantilla guardada en:\n{generada}\n\n"
            "Completala con tus maquinas y volve a este dialogo para importarla.",
        )

    def _on_preview(self) -> None:
        if not self._ruta_archivo:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self._ruta_archivo, data_only=True, read_only=True)
            # Preferimos hoja Master, si no la primera
            hoja = "Master" if "Master" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[hoja]
            filas = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            QMessageBox.critical(self, "Error al leer", f"No se pudo abrir el Excel:\n{e}")
            return

        if not filas:
            QMessageBox.warning(self, "Vacio", "El archivo no tiene filas.")
            return

        # Detectar donde arrancan los headers (fila 1 suele ser el header)
        headers_raw = filas[0]
        # Si la primera fila esta vacia, probar fila 2 o 3
        if not any(headers_raw):
            for offset in (1, 2):
                if len(filas) > offset and any(filas[offset]):
                    headers_raw = filas[offset]
                    filas = filas[offset + 1:]
                    break

        self._preview_headers = [
            str(h).strip() if h is not None else "" for h in headers_raw
        ]
        # Limitar columnas visibles a las 13 del Master (las que nos importan)
        # pero si el Excel tiene mas/menos columnas, respetamos lo que tenga
        self._preview_data = [
            tuple(self._safe(v) for v in fila)
            for fila in filas[:PREVIEW_ROWS]
            if any(v is not None and str(v).strip() for v in fila)
        ]

        self._poblar_tabla_preview()
        self._lbl_status.setText(
            f"Preview: hoja '{hoja}', {len(filas)} filas totales. "
            f"Mostrando las primeras {len(self._preview_data)} con datos."
        )
        self._btn_importar.setEnabled(True)

    def _safe(self, v) -> str:
        if v is None:
            return ""
        if isinstance(v, float) and v != v:
            return ""
        return str(v).strip()

    def _poblar_tabla_preview(self) -> None:
        cols = len(self._preview_headers)
        self._tabla.setColumnCount(cols)
        self._tabla.setHorizontalHeaderLabels(self._preview_headers)
        self._tabla.setRowCount(len(self._preview_data))
        for r_idx, fila in enumerate(self._preview_data):
            for c_idx in range(cols):
                valor = fila[c_idx] if c_idx < len(fila) else ""
                item = QTableWidgetItem(valor)
                self._tabla.setItem(r_idx, c_idx, item)
        # Ajustar columnas razonables
        header = self._tabla.horizontalHeader()
        for c in range(cols):
            header.setSectionResizeMode(c, QHeaderView.ResizeToContents)

    def _on_importar(self) -> None:
        if not self._ruta_archivo:
            return
        # Confirmacion
        res = QMessageBox.question(
            self,
            "Confirmar importacion",
            f"Importar maquinas desde:\n{self._ruta_archivo}\n\n"
            "Las maquinas existentes seran actualizadas; las nuevas se insertaran. "
            "Seguir?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if res != QMessageBox.Yes:
            return

        self._set_busy(True, "Importando...")
        # QApplication.processEvents() para refrescar UI antes del import largo
        from PySide6.QtWidgets import QApplication
        QApplication.processEvents()

        try:
            resultado = svc_maq.importar_desde_excel(self._ruta_archivo)
        except Exception as e:
            self._set_busy(False, "")
            QMessageBox.critical(self, "Error al importar", str(e))
            return
        finally:
            self._set_busy(False, "")

        self._lbl_status.setText(
            f"Listo: {resultado['insertadas']} insertadas, "
            f"{resultado['actualizadas']} actualizadas, "
            f"{len(resultado['errores'])} errores."
        )

        msg = (
            f"Hoja usada: {resultado['hoja_usada']}\\n"
            f"Insertadas: {resultado['insertadas']}\\n"
            f"Actualizadas: {resultado['actualizadas']}\\n"
            f"Errores: {len(resultado['errores'])}"
        )
        if resultado["errores"]:
            muestra = "\\n".join(resultado["errores"][:10])
            msg += f"\\n\\nPrimeros errores:\\n{muestra}"

        QMessageBox.information(self, "Resultado de importacion", msg)
        self.finished_with_result.emit(resultado)

    def _set_busy(self, busy: bool, mensaje: str) -> None:
        self._progress.setVisible(busy)
        self._btn_preview.setEnabled(not busy and self._ruta_archivo is not None)
        self._btn_importar.setEnabled(not busy and self._ruta_archivo is not None)
        self._btn_sel = self.findChild(QPushButton, "")  # noop
        if mensaje:
            self._lbl_status.setText(mensaje)


__all__ = ("ImportDialog",)